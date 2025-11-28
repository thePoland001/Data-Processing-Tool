from __future__ import annotations

import os
import json
import time
import pathlib
from typing import Dict, Any

import google.generativeai as genai
from jsonschema import validate, ValidationError
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from dateutil import parser as dateparser  # pip install python-dateutil

from ingest import load_eml_texts

# -----------------------------
# configuration
# -----------------------------
DESKTOP = pathlib.Path.home() / "Desktop"

CY_DIR = DESKTOP / "CY 25"
REJECTS_DIR = DESKTOP / "rejects"
OUT_XLSX = DESKTOP / "valid.xlsx"
SYSTEM_PROMPT_FILE = pathlib.Path("system_prompt.md")

COLUMNS = [
    "Date",
    "Location",
    "Details",
    "Category",
    "Report Type",
    "Operation/Activity",
    "Contributing Factors",
    "Notes",
]

SCHEMA: Dict[str, Any] = {
    "type": "object",
    "properties": {col: {"type": "string"} for col in COLUMNS},
    "required": COLUMNS,
    "additionalProperties": False,
}

# dropdown values
LOCATION_VALUES = ["COS", "FGA", "HSV", "VSFB", "SBX", "LRDR", "TBD"]
CATEGORY_VALUES = ["Environmental", "Hazard Alert", "Health", "Injury", "Near Miss", "Property Damage", "TBD"]
REPORT_TYPE_VALUES = [
    "Hazardous object",
    "Equipment failure",
    "Environment hazard",
    "Hazardous substance",
    "Unplanned release",
    "Slip or trip",
    "Transportation incident",
    "TBD/Other",
    "TBD",
]
OPERATION_VALUES = [
    "AMBULATORY",
    "ERGONOMIC",
    "HAZMAT HANDLING",
    "ELECTRICAL or HOT WORK",
    "INSPECTING/TESTING",
    "OPERATING PORTABLE TOOLS/APPLIANCES",
    "OPERATING MOBILE EQUIPMENT",
    "OPERATING STATIONARY EQUIPMENT/CRANES",
    "OFFICE/OTHER",
]

# api key (sample fallback)
GEMINI_API_KEY = os.environ.get("GEMINI_API_KEY") or "AIzaSyCrAIuCpLeYT3gBC9494GVb95ErXIg0NtQ"
MODEL_NAME = os.environ.get("GEMINI_MODEL", "gemini-1.5-flash")

genai.configure(api_key=GEMINI_API_KEY)
model = genai.GenerativeModel(MODEL_NAME)

# -----------------------------
# helpers (LLM + normalization)
# -----------------------------
def _read_system_prompt() -> str:
    if SYSTEM_PROMPT_FILE.exists():
        return SYSTEM_PROMPT_FILE.read_text(encoding="utf-8")
    return "Extract JSON with keys: " + ", ".join(COLUMNS) + '. Use dropdown values or "TBD".'

def _nudge_for_json(email_text: str) -> str:
    return (
        "Your previous response was not valid JSON. "
        "Return only valid JSON that conforms to the schema, nothing else.\n\n"
        "EMAIL TEXT:\n" + email_text
    )

def call_gemini_for_json(email_text: str, sys_prompt: str, retries: int = 3) -> Dict[str, Any]:
    prompt = f"""{sys_prompt}

EMAIL TEXT:
{email_text}
"""
    for attempt in range(1, retries + 1):
        resp = model.generate_content(prompt)
        text = (getattr(resp, "text", "") or "").strip()
        if text.startswith("```"):
            text = text.strip("`")
            if text.lower().startswith("json"):
                text = text[4:].strip()
        try:
            return json.loads(text)
        except Exception:
            prompt = _nudge_for_json(email_text)
            time.sleep(0.5 * attempt)
    raise RuntimeError("Gemini did not return valid JSON after retries")

# --- post-processing to guarantee filled fields ---
_KEYWORDS = {
    "AMBULATORY": ["walk", "step", "slip", "trip", "fall", "ladder", "stairs", "climb", "descending"],
    "ERGONOMIC": ["lift", "carri", "push", "pull", "strain", "manual", "posture", "overexert"],
    "HAZMAT HANDLING": ["chemical", "corrosive", "reactive", "toxic", "spill", "release", "container", "msds"],
    "ELECTRICAL or HOT WORK": ["electr", "wiring", "breaker", "480", "arc", "energized", "weld", "torch", "hot work"],
    "INSPECTING/TESTING": ["inspect", "test", "measurement", "verify", "calibrat"],
    "OPERATING PORTABLE TOOLS/APPLIANCES": ["drill", "grinder", "saw", "hand tool", "portable"],
    "OPERATING MOBILE EQUIPMENT": ["vehicle", "truck", "loader", "tractor", "forklift", "gsa vehicle", "mobile platform"],
    "OPERATING STATIONARY EQUIPMENT/CRANES": ["crane", "hoist", "lathe", "press", "compressor", "generator", "stationary"],
    "OFFICE/OTHER": ["office", "admin", "desk", "paperwork", "design"],
}

def normalize_mdy(s: str) -> str:
    try:
        dt = dateparser.parse(s, dayfirst=False, yearfirst=False, fuzzy=True)
        return f"{dt.month}/{dt.day}/{dt.year}"
    except Exception:
        return s

def infer_operation_activity(email_text: str) -> str:
    t = email_text.lower()
    scores = {k: 0 for k in _KEYWORDS}
    for k, terms in _KEYWORDS.items():
        for w in terms:
            if w in t:
                scores[k] += 1
    best = max(scores, key=scores.get)
    return best if scores[best] > 0 else "OFFICE/OTHER"

def postprocess(data: Dict[str, Any], email_text: str) -> Dict[str, Any]:
    # normalize date
    if "Date" in data:
        data["Date"] = normalize_mdy(str(data["Date"]))
    # operation/activity must never be TBD/empty
    op = (data.get("Operation/Activity") or "").strip()
    if not op or op.upper() == "TBD":
        data["Operation/Activity"] = infer_operation_activity(email_text)
    # notes must not be empty
    notes = (data.get("Notes") or "").strip()
    if not notes or notes.upper() == "TBD":
        details = (data.get("Details") or "").strip()
        snippet = email_text.strip().replace("\n", " ")[:220]
        data["Notes"] = details if details else snippet
    return data

def row_from_data(data: Dict[str, Any]) -> list[str]:
    validate(instance=data, schema=SCHEMA)
    return [str(data.get(col, "TBD") or "TBD") for col in COLUMNS]

# -----------------------------
# workbook setup + formatting
# -----------------------------
HEADER_FILL = PatternFill("solid", fgColor="E67E22")  # orange
ROW_FILL = PatternFill("solid", fgColor="DCEAF7")     # light blue
HEADER_FONT = Font(bold=True, color="000000")
CENTER = Alignment(vertical="center")
THIN = Side(style="thin", color="000000")
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COL_WIDTHS = [12, 12, 40, 18, 22, 28, 28, 48]  # A..H widths

def style_header(ws) -> None:
    for col_idx, title in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER_ALL
        # set column width
        ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS[col_idx - 1]
    # freeze header and enable filters
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

def style_row(ws, row_idx: int) -> None:
    # shade row and add borders
    for col_idx in range(1, len(COLUMNS) + 1):
        c = ws.cell(row=row_idx, column=col_idx)
        c.fill = ROW_FILL
        c.alignment = CENTER if col_idx in (1, 2, 4, 5, 6) else Alignment(wrap_text=True, vertical="top")
        c.border = BORDER_ALL

def setup_workbook(path: pathlib.Path) -> Workbook:
    max_rows = 5000
    if path.exists():
        wb = load_workbook(path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(COLUMNS)
        style_header(ws)

        # dropdowns
        dv_location = DataValidation(type="list", formula1=f'"{",".join(LOCATION_VALUES)}"', allow_blank=True)
        dv_category = DataValidation(type="list", formula1=f'"{",".join(CATEGORY_VALUES)}"', allow_blank=True)
        dv_report   = DataValidation(type="list", formula1=f'"{",".join(REPORT_TYPE_VALUES)}"', allow_blank=True)
        dv_operation= DataValidation(type="list", formula1=f'"{",".join(OPERATION_VALUES)}"', allow_blank=True)

        ws.add_data_validation(dv_location)
        ws.add_data_validation(dv_category)
        ws.add_data_validation(dv_report)
        ws.add_data_validation(dv_operation)

        # apply to columns (B, D, E, F)
        dv_location.add(f"B2:B{max_rows}")
        dv_category.add(f"D2:D{max_rows}")
        dv_report.add(f"E2:E{max_rows}")
        dv_operation.add(f"F2:F{max_rows}")

    return wb

# -----------------------------
# main (5-second watcher)
# -----------------------------
def main() -> None:
    CY_DIR.mkdir(exist_ok=True, parents=True)
    REJECTS_DIR.mkdir(exist_ok=True, parents=True)

    sys_prompt = _read_system_prompt()
    wb = setup_workbook(OUT_XLSX)
    ws = wb.active

    seen: set[str] = set()

    print("watching ~/Desktop/CY 25 every 5s... (Ctrl+C to stop)")
    while True:
        for eml_path, email_text in load_eml_texts(CY_DIR):
            if eml_path.name in seen:
                continue
            try:
                data = call_gemini_for_json(email_text, sys_prompt)
                data = postprocess(data, email_text)
                row = row_from_data(data)

                ws.append(row)
                # style the just-added row
                style_row(ws, ws.max_row)

                wb.save(OUT_XLSX)
                seen.add(eml_path.name)
                print(f"✔ processed {eml_path.name}")
            except (ValidationError, RuntimeError, Exception) as e:
                reject_path = REJECTS_DIR / f"{eml_path.stem}.reject.txt"
                reject_path.write_text(f"{type(e).__name__}: {e}\n", encoding="utf-8")
                seen.add(eml_path.name)
                print(f"✘ rejected {eml_path.name}: {e}")

        time.sleep(5)

if __name__ == "__main__":
    main()
